from openpyxl import Workbook, load_workbook
import re


class ymsports_summary_rank():
    def __init__(self, input_filename, output_filename, class_counts):
        self.input_filename = input_filename
        self.output_filename = output_filename
        self.class_counts = class_counts
        self.rank_number = 4
        self.sports_ranks = {}
        self.wb = load_workbook(
            filename=f'{self.input_filename}.xlsx', read_only=True)
        self.get_top_four()
        self.write_result_to_a_workbook()

    def get_top_four(self):
        sports_ranks_raw = {}
        for sheet in self.wb.sheetnames:
            self.class_count = self.class_counts[self.wb.sheetnames.index(
                sheet)]
            sports_ranks = self.get_sports_items_result(sheet)
            if sports_ranks_raw == {}:
                sports_ranks_raw = sports_ranks
            else:
                sports_ranks_raw = {
                    key: sports_ranks_raw[key] + sports_ranks[key] for key in sports_ranks_raw if key in sports_ranks}

        for key, project in sports_ranks_raw.items():
            if project[0][2] == None:
                continue

            if len(re.split('\'|"', project[0][2])) > 1:
                project_sorted = sorted(
                    project, key=lambda item: item[2])
            else:
                project_sorted = sorted(
                    project, key=lambda item: item[2], reverse=True)
            self.sports_ranks[key] = project_sorted[0:4]

    # read data from original spreadsheet and return a data dictionary

    def get_sports_items_result(self, sheet_name):
        ws = self.wb[sheet_name]

        # sports ranks
        sports_ranks = {}

        for i in range(24):

            single_row = ws[5+i]
            sports_item = single_row[0].value
            sports_ranks[sports_item] = []

            for i in range(self.rank_number):
                rows = single_row[i*4+3: i*4+7]

                if rows[1].value != None:
                    name = rows[0].value
                    class_no = sheet_name[0] + '-' + rows[1].value
                    grade = rows[2].value
                    sports_ranks[sports_item].append([name, class_no, grade])

        return sports_ranks

    # use the given dictionary data to write to a new spread sheeet
    def write_result_to_a_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'rank_school'

        header_names = ['项目'] + ['姓名', '班级', '最好成绩', '名次']*4

        for header in range(len(header_names)):
            ws.cell(1, header+1, header_names[header])

        #     ws.cell(sider+2, 1, f'{self.output_filename[:5]}级{sider+1:02}班')

        # # `k` control how the columns will be set
        k = 1
        for rank_name, rank_items in self.sports_ranks.items():
            ws.cell(k+1, 1,  rank_name)
            for position, rank_item in enumerate(rank_items):
                # print(rank_item)
                # value = sports_items[sports_item][i]
                name = rank_item[0]
                class_code = re.split('-', rank_item[1])
                class_code = f'{class_code[0]}年级{int(class_code[1]):02}班'
                decent_score = rank_item[2]
                rank_number = position+1
                # print(name, class_code, decent_score, rank_name)

                ws.cell(k+1, position*4+2, name)
                ws.cell(k+1, position*4+3, class_code)
                ws.cell(k+1, position*4+4, decent_score)
                ws.cell(k+1, position*4+5, rank_number)

            k += 1
        # Save it to a xlsx spreadsheet
        wb.save(f'{self.output_filename}.xlsx')
