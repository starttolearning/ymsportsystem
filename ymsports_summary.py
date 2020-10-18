from openpyxl import Workbook, load_workbook


class ymsports_summary():
    def __init__(self, input_filename, output_filename, class_counts):
        self.input_filename = input_filename
        self.output_filename = output_filename+'.xlsx'
        self.class_counts = class_counts
        self.wb = load_workbook(
            filename=f'{self.input_filename}.xlsx', read_only=True)
        self.wbw = load_workbook(filename=self.output_filename)

        for sheet in self.wb.sheetnames:
            self.class_count = self.class_counts[self.wb.sheetnames.index(
                sheet)]
            sports_items = self.get_sports_items_result(sheet)
            # self.write_result_to_a_workbook(sports_items, sheet)
            self.write_result_to_a_exist_workbook(sports_items, sheet)

        self.wbw.save(self.output_filename)

    # read data from original spreadsheet and return a data dictionary

    def get_sports_items_result(self, sheet_name):
        ws = self.wb[sheet_name]
        sports_items = {}

        for i in range(24):

            single_row = ws[5+i]
            sports_item = single_row[0].value
            class_values = {k: 0 for k in range(1, self.class_count+1)}

            for i in range(8):
                rows = single_row[i*4+3: i*4+7]
                if rows[1].value != None:
                    class_values[int(rows[1].value)] += int(rows[3].value)

            sports_items[sports_item] = class_values
        return sports_items

    # use the given dictionary data to write to a new spread sheeet
    def write_result_to_a_workbook(self, sports_items, sheet_name):
        wb = Workbook()
        ws = wb.active
        # ws.title = 'sh'
        sports_items_name = list(sports_items.keys())

        for header in range(len(sports_items)):
            ws.cell(1, header+2, sports_items_name[header])

        for sider in range(self.class_count):
            ws.cell(sider+2, 1, f'{self.output_filename[:4]}级{sider+1:02}班')

        # `k` control how the columns will be set
        k = 1
        for sports_item in sports_items:
            for i in sports_items[sports_item]:
                value = sports_items[sports_item][i]
                if value == 0:
                    ws.cell(i+1, k+1, '-')
                else:
                    ws.cell(i+1, k+1, value)
            k += 1
        # Save it to a xlsx spreadsheet
        wb.save(f'{self.output_filename}-{sheet_name}.xlsx')

    # use the given dictionary data to write to a new spread sheeet
    def write_result_to_a_exist_workbook(self, sports_items, sheet_name):
        ws = self.wbw[sheet_name]
        # print(sports_items)
        # ws.title = 'sh'

        # `k` control how the columns will be set
        k = 1
        for sports_item in sports_items:
            for i in sports_items[sports_item]:
                value = sports_items[sports_item][i]
                if value == 0:
                    ws.cell(i+3, k+1, '-')
                else:
                    ws.cell(i+3, k+1, value)
            k += 1
        # Save it to a xlsx spreadsheet
